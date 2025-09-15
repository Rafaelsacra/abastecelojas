import openpyxl
import math
import os
import re
from openpyxl.styles import PatternFill

# === Configuração das lojas ===
LOJAS = {
    "02 JANAUBA": r"\\192.168.1.185\Compras\ANDERSON - JOSIANE\PEDIDO LOJA\METOD NOVA\02 JANAUBA ABASTECIMENTO PADRAO.xlsx",
    "04 MAJOR PRATES": r"\\192.168.1.185\Compras\ANDERSON - JOSIANE\PEDIDO LOJA\METOD NOVA\04 MAJOR PRATES ABASTECIMENTO PADRAO.xlsx",
    "05 SAO JOSE": r"\\192.168.1.185\Compras\ANDERSON - JOSIANE\PEDIDO LOJA\METOD NOVA\05 SAO JOSE ABASTECIMENTO PADRAO.xlsx",
    "07 SALINAS": r"C:\Users\User\Documents\DESENVOLVIMENTO WEB\TESTE\07 SALINAS ABASTECIMENTO PADRAO.xlsx",
    "08 PIRAPORA": r"\\192.168.1.185\Compras\ANDERSON - JOSIANE\PEDIDO LOJA\METOD NOVA\08 PIRAPORA ABASTECIMENTO PADRAO.xlsx",
}

# === Colunas fixas ===
COLUNAS = {
    "embalagem": 3,      # C
    "estoque_matriz": 4, # D
    "dias_matriz": 5,    # E
    "estoque_loja": 24,  # X
    "dias_loja": 25,     # Y
    "pedir": 26,         # Z
    "regra_aplicada": 29,  # AC
    "media_venda": 32,   # AF
    "vendas_historico": list(range(15, 23)),  # Colunas O-W (15-23)
}

# === Cores para as regras ===
CORES = {
    "verde": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Verde claro
    "laranja": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),    # Laranja
    "rosa": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),       # Rosa claro
    "amarelo": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),    # Amarelo claro
    "vermelho_claro": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"), # Vermelho claro
}

# === Funções auxiliares ===
def extrair_numero_embalagem(embalagem_str):
    """Extrai o número da embalagem (ex: CX6 → 6, EV10 → 10, KG20 → 20)"""
    if not embalagem_str:
        return 1
    
    # Converter para string e remover espaços
    embalagem_str = str(embalagem_str).strip()
    
    # Extrair apenas os números
    numeros = re.findall(r'\d+', embalagem_str)
    
    if numeros:
        return int(numeros[0])
    else:
        return 1  # Padrão se não encontrar número

def arredondar_para_multiplo(pedir_bruto, embalagem, estoque_loja_atual=0, estoque_matriz_atual=0, media_venda=0):
    """Arredonda pedir_bruto para múltiplo da embalagem"""
    if pedir_bruto <= 0:
        return 0
    
    return math.ceil(pedir_bruto / embalagem) * embalagem

def calcular_vendas_historico(ws, row):
    """Calcula o total de vendas do histórico (colunas O-W)"""
    total_vendas = 0
    for col in COLUNAS["vendas_historico"]:
        try:
            valor = ws.cell(row=row, column=col).value or 0
            total_vendas += float(valor)
        except:
            pass
    return total_vendas

def buscar_maior_valor_historico(ws, row):
    """Busca o maior valor nas colunas O-W"""
    maior_valor = 0
    for col in COLUNAS["vendas_historico"]:
        try:
            valor = ws.cell(row=row, column=col).value or 0
            maior_valor = max(maior_valor, float(valor))
        except:
            pass
    return maior_valor

def calcular_pedido_baseado_vendas(ws, row, embalagem):
    """Calcula pedido baseado nas vendas históricas"""
    vendas_historico = calcular_vendas_historico(ws, row)
    if vendas_historico > 0:
        # Média de vendas por mês
        media_mensal = vendas_historico / len(COLUNAS["vendas_historico"])
        # Sugestão para 30 dias
        sugestao = media_mensal
        # Ajustar pela embalagem
        if sugestao > 0:
            return arredondar_para_multiplo(sugestao, embalagem)
    return 0

def aplicar_cor_celula(ws, row, col, cor):
    """Aplica cor à célula"""
    ws.cell(row=row, column=col).fill = CORES[cor]

# === Pergunta dias de cobertura ===
dias_cobertura_frios = int(input("Informe os dias de cobertura para FRIOS: "))
dias_cobertura_condimentos = int(input("Informe os dias de cobertura para CONDIMENTOS: "))

# === Pergunta lojas ===
print("\nSelecione as lojas para analisar (separe por vírgula):")
for i, loja in enumerate(LOJAS.keys(), start=1):
    print(f"{i} - {loja}")

opcoes = input("Opção(s): ").split(",")
selecionadas = [list(LOJAS.keys())[int(i.strip()) - 1] for i in opcoes]

for loja in selecionadas:
    caminho = LOJAS[loja]

    if not os.path.exists(caminho):
        print(f"❌ Arquivo não encontrado: {caminho}")
        continue

    print(f"\n📂 Processando {loja}...")

    # === Abrir planilha com valores congelados das fórmulas ===
    wb = openpyxl.load_workbook(caminho, data_only=True)

    # Criar/limpar aba de relatório
    if "RELATORIO" in wb.sheetnames:
        ws_rel = wb["RELATORIO"]
        wb.remove(ws_rel)
    ws_rel = wb.create_sheet("RELATORIO")
    ws_rel.append(["ABA", "CODIGO", "DESCRICAO", "PEDIR", "REGRA_APLICADA", "VENDA_DIARIA"])
    
    # Adicionar legenda das cores
    ws_rel.append([])
    ws_rel.append(["FLUXO COMPLETO DE 4 ETAPAS:"])
    ws_rel.append(["1ª PARTE - Regras iniciais:"])
    ws_rel.append(["  VERDE", "Estoque Ideal com Embalagem"])
    ws_rel.append(["  BRANCO", "Z = Estoque Matriz Zerado | R = Estoque Matriz Baixo"])
    ws_rel.append(["2ª PARTE - Ajuste Matriz:"])
    ws_rel.append(["  LARANJA", "Ajustado para média das vendas mensais"])
    ws_rel.append(["3ª PARTE - Regras complementares:"])
    ws_rel.append(["  ROSA", "Estoque Loja Zerado + Sem Vendas + Maior Histórico (ou 1 se histórico=0)"])
    ws_rel.append(["  AMARELO", "Dias de Estoque Baixo + Maior Histórico"])
    ws_rel.append(["4ª PARTE - Reajuste Matriz:"])
    ws_rel.append(["  VERMELHO CLARO", "Dias > 10: Zerado | Dias < 10: Completar até 10 dias"])
    ws_rel.append([])
    ws_rel.append(["CONFIGURAÇÕES:"])
    ws_rel.append(["• Sempre múltiplo da embalagem"])
    ws_rel.append(["• Arredondamento para cima"])
    ws_rel.append(["• Aplicação sequencial das 4 etapas"])
    ws_rel.append([])
    ws_rel.append(["ABA", "CODIGO", "DESCRICAO", "PEDIR", "REGRA_APLICADA", "VENDA_DIARIA"])

    # === Processar abas Frios e Condimentos ===
    for aba in ["FRIOS", "CONDIMENTOS"]:
        if aba not in wb.sheetnames:
            print(f"⚠ Aba {aba} não encontrada, pulando...")
            continue

        ws = wb[aba]
        dias_cobertura = dias_cobertura_frios if aba == "FRIOS" else dias_cobertura_condimentos

        for row in range(3, ws.max_row + 1):  # começa na linha 3
            try:
                # Ler dados das células
                embalagem_str = ws.cell(row=row, column=COLUNAS["embalagem"]).value or ""
                estoque_matriz = ws.cell(row=row, column=COLUNAS["estoque_matriz"]).value or 0
                dias_matriz = ws.cell(row=row, column=COLUNAS["dias_matriz"]).value or 0
                estoque_loja = ws.cell(row=row, column=COLUNAS["estoque_loja"]).value or 0
                dias_loja = ws.cell(row=row, column=COLUNAS["dias_loja"]).value or 0
                media_venda = ws.cell(row=row, column=COLUNAS["media_venda"]).value or 0

                # Extrair número da embalagem (ex: CX6 → 6)
                embalagem = extrair_numero_embalagem(embalagem_str)
                try:
                    estoque_matriz = float(estoque_matriz) if estoque_matriz else 0
                except:
                    estoque_matriz = 0
                try:
                    dias_matriz = float(dias_matriz) if dias_matriz else 0
                except:
                    dias_matriz = 0
                try:
                    estoque_loja = float(estoque_loja) if estoque_loja else 0
                except:
                    estoque_loja = 0
                try:
                    dias_loja = float(dias_loja) if dias_loja else 0
                except:
                    dias_loja = 0
                try:
                    media_venda = float(media_venda) if media_venda else 0
                except:
                    media_venda = 0

                # Verificar se linha está vazia
                if not any([embalagem_str, estoque_matriz, estoque_loja, media_venda]):
                    continue

                # === FLUXO COMPLETO DE 4 ETAPAS ===
                
                # Inicializar variáveis
                pedir_final = ""
                cor_aplicada = None
                regra_aplicada = ""
                
                # === 1ª PARTE: Regras iniciais de preenchimento ===
                
                # Zerado na Matriz (D = 0) → PEDIR (Z) = "Z"
                if estoque_matriz <= 0:
                    pedir_final = "Z"
                    regra_aplicada = "Zerado Matriz"
                
                # Estoque baixo na Matriz (1 ≤ E ≤ 10) → PEDIR (Z) = "R"
                elif 1 <= dias_matriz <= 10:
                    pedir_final = "R"
                    regra_aplicada = "Baixo Estoque Matriz"
                
                # Caso contrário: calcular estoque ideal
                else:
                    demanda_dia = media_venda / 30 if media_venda > 0 else 0
                    estoque_ideal = demanda_dia * dias_cobertura
                    pedir_bruto = estoque_ideal - estoque_loja
                    
                    if pedir_bruto > 0:
                        pedir_final = arredondar_para_multiplo(pedir_bruto, embalagem)
                        cor_aplicada = "verde"
                        regra_aplicada = "Estoque Ideal"
                
                # === 2ª PARTE: Ajuste pelo estoque da Matriz ===
                
                # Para cada linha com PEDIR (Z) numérico
                if isinstance(pedir_final, (int, float)) and pedir_final > 0:
                    limite_30_matriz = estoque_matriz * 0.30
                    
                    if pedir_final > limite_30_matriz:
                        # Calcular média das vendas mensais (colunas O-W)
                        vendas_historico = calcular_vendas_historico(ws, row)
                        media_mensal = vendas_historico / len(COLUNAS["vendas_historico"]) if vendas_historico > 0 else 0
                        
                        if media_mensal > 0:
                            pedir_final = round(media_mensal)  # Usar média das vendas mensais
                            cor_aplicada = "laranja"
                            regra_aplicada = "Ajuste 30% Média Vendas"
                        else:
                            # Se não há vendas históricas, manter o valor original
                            pedir_final = pedir_final
                            cor_aplicada = "laranja"
                            regra_aplicada = "Ajuste 30% Sem Vendas"
                
                # === 3ª PARTE: Regras complementares ===
                
                # Estoque Loja Zerado + Média de Vendas = 0 + PEDIR = 0
                if (estoque_loja <= 0 and media_venda == 0 and 
                    (pedir_final == "" or pedir_final == 0)):
                    
                    maior_valor = buscar_maior_valor_historico(ws, row)
                    if maior_valor > 0:
                        pedir_final = arredondar_para_multiplo(maior_valor, embalagem)
                        cor_aplicada = "rosa"
                        regra_aplicada = "Estoque Loja Zerado + Vendas Zeradas"
                    else:
                        # Se o maior valor for 0, sugerir 1
                        pedir_final = 1
                        cor_aplicada = "rosa"
                        regra_aplicada = "Estoque Loja Zerado + Vendas Zeradas"
                
                # Dias de Estoque Loja Baixo (1 ≤ Y ≤ 12) + PEDIR = 0
                if (1 <= dias_loja <= 12 and 
                    (pedir_final == "" or pedir_final == 0)):
                    
                    maior_valor = buscar_maior_valor_historico(ws, row)
                    if maior_valor > 0:
                        # Subtrair o estoque atual da loja
                        pedir_bruto = maior_valor - estoque_loja
                    if pedir_bruto > 0:
                        pedir_final = arredondar_para_multiplo(pedir_bruto, embalagem)
                        cor_aplicada = "amarelo"
                        regra_aplicada = "Dias Estoque Loja Baixo"
                
                # === 4ª PARTE: Reaplicar Ajuste pelo estoque da Matriz ===
                
                # Repetir validação do limite de 30% da Matriz
                if isinstance(pedir_final, (int, float)) and pedir_final > 0:
                    limite_30_matriz = estoque_matriz * 0.30
                    
                    if pedir_final > limite_30_matriz:
                        # Avaliar estoque Loja (X) e dias de estoque Loja (Y)
                        if dias_loja > 10:
                            # Dias de cobertura acima de 10 - não sugerir pedido
                            pedir_final = 0
                            cor_aplicada = "vermelho_claro"
                            regra_aplicada = "Reaplicar 30% Cobertura > 10 dias"
                        else:
                            # Dias de cobertura insuficiente - completar até 10 dias
                            if media_venda > 0:
                                # Calcular quantidade necessária para 10 dias
                                demanda_diaria = media_venda / 30
                                estoque_ideal_10_dias = demanda_diaria * 10
                                quantidade_necessaria = estoque_ideal_10_dias - estoque_loja
                                
                                if quantidade_necessaria > 0:
                                    pedir_final = arredondar_para_multiplo(quantidade_necessaria, embalagem)
                                    cor_aplicada = "vermelho_claro"
                                    regra_aplicada = "Reaplicar 30% Cobertura < 10 dias"
                                else:
                                    pedir_final = 0
                                    cor_aplicada = "vermelho_claro"
                                    regra_aplicada = "Reaplicar 30% Cobertura < 10 dias"
                            else:
                                # Sem vendas históricas, manter valor original limitado
                                pedir_final = round(limite_30_matriz)
                                cor_aplicada = "vermelho_claro"
                                regra_aplicada = "Reaplicar 30% Sem Vendas"

                # Gravar resultado na célula
                ws.cell(row=row, column=COLUNAS["pedir"], value=pedir_final)
                
                # Gravar regra aplicada na coluna AC
                ws.cell(row=row, column=COLUNAS["regra_aplicada"], value=regra_aplicada)
                
                # Aplicar cor se necessário
                if cor_aplicada:
                    aplicar_cor_celula(ws, row, COLUNAS["pedir"], cor_aplicada)

                # Adicionar ao relatório
                codigo = ws.cell(row=row, column=1).value
                descricao = ws.cell(row=row, column=2).value
                venda_diaria = round(media_venda / 30, 2) if media_venda > 0 else 0
                ws_rel.append([aba, codigo, descricao, pedir_final, regra_aplicada, venda_diaria])

            except Exception as e:
                print(f"⚠ Erro na linha {row} da aba {aba}: {e}")

    # Salvar
    wb.save(caminho)
    print(f"✅ Arquivo atualizado: {caminho}")


